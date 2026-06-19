"""Multi-catalog loader."""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Dict, Optional

from app.services.linework_grammar import GRAMMARS, LineworkGrammar

import logging, sys

log = logging.getLogger("robs_code_wizard")


@dataclass
class CodeSet:
    name: str
    codes: List[Dict[str, str]] = field(default_factory=list)
    linework_grammar: Optional[LineworkGrammar] = None
    parser_kind: str = "pnezd"
    source_path: Optional[Path] = None

    @property
    def code_set(self) -> frozenset:
        return frozenset(c["code"].upper() for c in self.codes if c.get("code"))

    def is_known_code(self, token: str) -> bool:
        if not token:
            return False
        return token.upper() in self.code_set


CATALOG_FILES = {"vdt": "VDT_CODES.xlsx", "odot": "ODOT_CODES.xlsx"}
PARSER_KIND = {"vdt": "pnezd", "odot": "odot"}


def _resources_dir() -> Path:
    """Locate resources/ for both dev and PyInstaller-frozen runs."""
    # PyInstaller --onefile extracts to sys._MEIPASS; the spec --add-data
    # places resources/ there at the top level.
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        bundled = Path(meipass) / "resources"
        if bundled.exists():
            return bundled
    # Dev / source-tree fallback: <project_root>/resources next to app/
    return Path(__file__).resolve().parents[2] / "resources"


def load_catalog(name: str) -> CodeSet:
    key = name.strip().lower()
    if key not in CATALOG_FILES:
        raise ValueError(f"Unknown catalog name: {name!r}; expected 'vdt' or 'odot'")

    grammar = GRAMMARS[key]
    parser_kind = PARSER_KIND[key]
    xlsx = _resources_dir() / CATALOG_FILES[key]
    if not xlsx.exists():
        log.warning("load_catalog(%s): %s not found, returning empty codeset", key, xlsx)
        return CodeSet(name=key, linework_grammar=grammar,
                       parser_kind=parser_kind, source_path=xlsx)

    try:
        from openpyxl import load_workbook
    except ImportError:
        log.warning("load_catalog(%s): openpyxl not installed, returning empty codeset", key)
        return CodeSet(name=key, linework_grammar=grammar,
                       parser_kind=parser_kind, source_path=xlsx)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, r in enumerate(rows):
        if r is None:
            continue
        lower = [str(c).strip().lower() if c is not None else "" for c in r]
        if "category" in lower and "code" in lower:
            header_idx = i
            break
    if header_idx is None:
        log.warning("load_catalog(%s): header row not found in %s, returning empty codeset", key, xlsx)
        return CodeSet(name=key, linework_grammar=grammar,
                       parser_kind=parser_kind, source_path=xlsx)

    header = [str(h).strip().lower() if h is not None else "" for h in rows[header_idx]]

    codes: List[Dict[str, str]] = []
    for raw in rows[header_idx + 1:]:
        if raw is None or all(c is None for c in raw):
            continue
        entry = dict(zip(header, [("" if c is None else str(c)) for c in raw]))
        code = entry.get("code", "").strip()
        if not code:
            continue
        codes.append({
            "category": entry.get("category", "").strip(),
            "code": code,
            "type": entry.get("type", "").strip(),
            "description": entry.get("description", "").strip(),
            "dtm": entry.get("dtm", "").strip(),
            "attributes_schema": entry.get("attributesschema", "").strip(),
            "symbol_cell": entry.get("symbolcell", "").strip(),
            "shot_location": entry.get("shotlocation", "").strip(),
        })
    log.info("load_catalog(%s) loaded %d codes from %s", key, len(codes), xlsx)
    return CodeSet(name=key, codes=codes, linework_grammar=grammar,
                   parser_kind=parser_kind, source_path=xlsx)
