"""Grammar-driven linework balance auditor (Phase 4 enhanced).

Adds row_indices tracking for the UI "Go to row" feature, a human-readable
suggestion engine, and TXT / XLSX export of the audit.
"""
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path


def _coerce_grammar(arg):
    if arg is None:
        return None
    if hasattr(arg, "linework_grammar"):
        return arg.linework_grammar
    return arg


def audit_with_locations(rows, grammar_or_codeset=None):
    """Audit linework balance and track where each command appeared.

    Returns a dict with: begins, ends, unbalanced, total_commands, row_indices.
    row_indices maps each code to the list of row indices where any begin or
    end command was applied to it (used by the UI "Go to row" buttons).
    """
    grammar = _coerce_grammar(grammar_or_codeset)
    if grammar is None:
        return {"begins": {}, "ends": {}, "unbalanced": {},
                "total_commands": 0, "row_indices": {}}

    begin_tokens = {t.upper() for t in (grammar.begin_line, grammar.begin_curve) if t}
    end_tokens = {t.upper() for t in (grammar.end_line, grammar.end_curve) if t}
    toggle = grammar.curve_toggle.upper() if grammar.curve_toggle else None

    begins, ends = Counter(), Counter()
    total = 0
    toggle_state = {}
    row_indices = defaultdict(list)

    for row_idx, row in enumerate(rows):
        desc = str(row.get("D", "")).strip().upper()
        if not desc:
            continue
        active = None
        for tok in desc.split():
            if tok in begin_tokens:
                if active:
                    begins[active] += 1
                    total += 1
                    row_indices[active].append(row_idx)
            elif tok in end_tokens:
                if active:
                    ends[active] += 1
                    total += 1
                    row_indices[active].append(row_idx)
            elif toggle and tok == toggle:
                if active:
                    state = toggle_state.get(active, "begin")
                    if state == "begin":
                        begins[active] += 1
                        toggle_state[active] = "end"
                    else:
                        ends[active] += 1
                        toggle_state[active] = "begin"
                    total += 1
                    row_indices[active].append(row_idx)
            else:
                active = tok

    unbalanced = {}
    for code in set(begins) | set(ends):
        d = begins[code] - ends[code]
        if d != 0:
            unbalanced[code] = d

    return {
        "begins": dict(begins),
        "ends": dict(ends),
        "unbalanced": unbalanced,
        "total_commands": total,
        "row_indices": {k: list(v) for k, v in row_indices.items()},
    }


def run_linework_fix(rows, grammar_or_codeset=None):
    """Backward-compatible wrapper around audit_with_locations.

    Returns the v0.3.9.3 shape (without row_indices) so existing callers and
    tests do not need to change.
    """
    a = audit_with_locations(rows, grammar_or_codeset)
    return {
        "begins": a["begins"],
        "ends": a["ends"],
        "unbalanced": a["unbalanced"],
        "total_commands": a["total_commands"],
    }


def suggest_fix(code, delta, grammar_or_codeset=None):
    """Return a human-readable fix suggestion for an unbalanced ``code``.

    ``delta > 0`` means there are more begin commands than end commands
    (linework was never closed).  ``delta < 0`` means there is an orphan end.
    """
    grammar = _coerce_grammar(grammar_or_codeset)
    if delta == 0:
        return f"{code}: balanced."
    if grammar is None:
        return f"{code}: unbalanced by {delta:+d} (no grammar to suggest a fix)."

    if delta > 0:
        end_tok = grammar.end_line or "<end>"
        n = delta
        plural = "s" if n != 1 else ""
        return (
            f"Missing {n} end command{plural} for {code}. "
            f"Add `{end_tok}` to the last `{code}` occurrence."
        )
    begin_tok = grammar.begin_line or "<begin>"
    n = -delta
    plural = "s" if n != 1 else ""
    return (
        f"Missing {n} begin command{plural} for {code}. "
        f"Add `{begin_tok}` to the first `{code}` occurrence."
    )


def export_audit_txt(audit, out_path):
    """Write a plain-text audit report.  Returns the output Path."""
    lines = [
        "Rob's Code Wizard - Linework Fix Audit",
        "=" * 40,
        f"Generated at      : {datetime.now().isoformat(timespec='seconds')}",
        f"Total commands    : {audit.get('total_commands', 0)}",
        f"Codes with begins : {len(audit.get('begins', {}))}",
        f"Codes with ends   : {len(audit.get('ends', {}))}",
        f"Unbalanced codes  : {len(audit.get('unbalanced', {}))}",
        "",
        "Per-Code Summary",
        "-" * 40,
        f"{'Code':<20}{'Begins':>8}{'Ends':>8}{'Delta':>8}",
    ]
    all_codes = sorted(set(audit.get("begins", {})) | set(audit.get("ends", {})))
    for code in all_codes:
        b = audit.get("begins", {}).get(code, 0)
        e = audit.get("ends", {}).get(code, 0)
        lines.append(f"{code:<20}{b:>8}{e:>8}{b-e:>+8}")

    unbal = audit.get("unbalanced", {})
    if unbal:
        lines.extend(["", "Unbalanced Detail", "-" * 40])
        for code, delta in unbal.items():
            lines.append(f"  {code}: {delta:+d}")

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(lines), encoding="utf-8")
    return out


def export_audit_xlsx(audit, out_path, grammar_or_codeset=None):
    """Write a multi-sheet xlsx audit report.  Returns the output Path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    grammar = _coerce_grammar(grammar_or_codeset)
    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    red = PatternFill("solid", fgColor="FFC7CE")
    hdr = PatternFill("solid", fgColor="E0E0E0")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    for k, v in [
        ("Report", "Rob's Code Wizard - Linework Fix Audit"),
        ("Generated at", datetime.now().isoformat(timespec="seconds")),
        ("Total commands", audit.get("total_commands", 0)),
        ("Codes with begins", len(audit.get("begins", {}))),
        ("Codes with ends", len(audit.get("ends", {}))),
        ("Unbalanced codes", len(audit.get("unbalanced", {}))),
    ]:
        ws.append([k, v])
    for cell in ws["A"]:
        cell.font = bold
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50

    ws2 = wb.create_sheet("Per-Code")
    headers = ["Code", "Begins", "Ends", "Delta", "Suggestion"]
    ws2.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = bold
        cell.alignment = center
        cell.fill = hdr

    all_codes = sorted(set(audit.get("begins", {})) | set(audit.get("ends", {})))
    for code in all_codes:
        b = audit.get("begins", {}).get(code, 0)
        e = audit.get("ends", {}).get(code, 0)
        d = b - e
        s = suggest_fix(code, d, grammar) if d != 0 else ""
        ws2.append([code, b, e, d, s])
        if d != 0:
            r = ws2.max_row
            for c in range(1, len(headers) + 1):
                ws2.cell(row=r, column=c).fill = red

    for col, w in zip("ABCDE", [16, 8, 8, 8, 70]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out