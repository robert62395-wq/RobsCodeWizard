"""Validation Report exporter."""
from collections import Counter
from datetime import datetime
from pathlib import Path
from app import __version__


def build_report(rows, results, suggestions, source_path=None):
    total = len(rows)
    valid_rows = sum(1 for r in results if r["valid"])
    invalid_rows = total - valid_rows
    zero_elev = 0
    for row in rows:
        try:
            if float(row.get("Z", 0) or 0) == 0.0:
                zero_elev += 1
        except (TypeError, ValueError):
            zero_elev += 1
    rows_with_suggestions = sum(1 for s in suggestions if s)
    ic = Counter()
    for res in results:
        for issue in res.get("issues", []):
            ic[issue.split(":")[0].strip()] += 1
    per_row = []
    for i, row in enumerate(rows):
        per_row.append({"point": row.get("P", ""), "original": row.get("D", ""),
                        "edited": row.get("D", ""), "valid": results[i]["valid"],
                        "issues": "; ".join(results[i].get("issues", [])),
                        "suggestion": suggestions[i] if i < len(suggestions) else ""})
    return {"summary": {"total_rows": total, "valid_rows": valid_rows,
                        "invalid_rows": invalid_rows, "zero_elevation_rows": zero_elev,
                        "rows_with_suggestions": rows_with_suggestions,
                        "unique_issue_counts": dict(ic)},
            "source_file": str(source_path) if source_path else "",
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "app_version": __version__, "per_row": per_row}


def export_xlsx(report, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "Summary"
    bold = Font(bold=True); center = Alignment(horizontal="center")
    s = report["summary"]
    for k, v in [("Report","Rob\'s Code Wizard Validation Report"),
                 ("App version",report["app_version"]),
                 ("Generated at",report["generated_at"]),
                 ("Source file",report["source_file"]),("",""),
                 ("Total rows",s["total_rows"]),("Valid rows",s["valid_rows"]),
                 ("Invalid rows",s["invalid_rows"]),("Zero elevation",s["zero_elevation_rows"]),
                 ("Rows w/ suggestion",s["rows_with_suggestions"])]:
        ws.append([k, v])
    for cell in ws["A"]: cell.font = bold
    ws.append([]); ws.append(["Issue Counts", ""])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["Issue", "Count"])
    hr = ws.max_row
    for c in (1,2):
        ws.cell(row=hr, column=c).font = bold; ws.cell(row=hr, column=c).alignment = center
    for label, count in s["unique_issue_counts"].items(): ws.append([label, count])
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 48
    ws2 = wb.create_sheet("Rows")
    headers = ["Point","Original","Edited","Valid","Issues","Suggestion"]
    ws2.append(headers)
    for c in range(1, len(headers)+1):
        ws2.cell(row=1, column=c).font = bold; ws2.cell(row=1, column=c).alignment = center
    for r in report["per_row"]:
        ws2.append([r["point"], r["original"], r["edited"],
                    "Yes" if r["valid"] else "No", r["issues"], r["suggestion"]])
    for i, w in enumerate([10,22,22,8,40,22], start=1):
        ws2.column_dimensions[chr(64+i)].width = w
    out_path = Path(out_path); out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path); return out_path


def export_txt(report, out_path):
    s = report["summary"]
    lines = ["Rob\'s Code Wizard Validation Report", "="*40,
             f"App version   : {report['app_version']}",
             f"Generated at  : {report['generated_at']}",
             f"Source file   : {report['source_file']}", "",
             f"Total rows         : {s['total_rows']}",
             f"Valid rows         : {s['valid_rows']}",
             f"Invalid rows       : {s['invalid_rows']}",
             f"Zero elevation     : {s['zero_elevation_rows']}",
             f"Rows w/ suggestion : {s['rows_with_suggestions']}", "",
             "Issue counts", "-"*40]
    for label, count in s["unique_issue_counts"].items():
        lines.append(f"  {label:<32} {count}")
    out_path = Path(out_path); out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8"); return out_path


def export_report(rows, results, suggestions, out_path, source_path=None):
    report = build_report(rows, results, suggestions, source_path=source_path)
    ext = Path(out_path).suffix.lower()
    if ext == ".xlsx": return export_xlsx(report, out_path)
    return export_txt(report, out_path)
