"""v0.4.4.1 hotfix: add PL (Overhead Powerline) to ODOT_CODES.xlsx if missing.

Idempotent - safe to run multiple times.
"""
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed. Run:")
    print("    .venv\\Scripts\\python -m pip install openpyxl")
    sys.exit(1)

PROJECT_ROOT = Path(__file__).parent
CANDIDATES = [
    PROJECT_ROOT / "app" / "data" / "ODOT_CODES.xlsx",
    PROJECT_ROOT / "resources" / "ODOT_CODES.xlsx",
    PROJECT_ROOT / "ODOT_CODES.xlsx",
]

NEW_CODE = "PL"
NEW_TYPE = "Linework"
NEW_DESC = "Overhead Powerline"


def find_xlsx():
    for p in CANDIDATES:
        if p.exists():
            print(f"[OK] Found ODOT_CODES.xlsx at: {p}")
            return p
    print("[ERROR] Could not find ODOT_CODES.xlsx. Looked in:")
    for p in CANDIDATES:
        print(f"    {p}")
    sys.exit(1)


def detect_header_row(ws):
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        lower = [str(c or "").strip().lower() for c in row]
        non_empty = sum(1 for c in lower if c)
        if non_empty < 3:
            continue
        cols = {}
        for i, h in enumerate(lower):
            if not h:
                continue
            if "code" in h and "code" not in cols:
                cols["code"] = i + 1
            elif ("type" in h or "category" in h or "class" in h) and "type" not in cols:
                cols["type"] = i + 1
            elif ("description" in h or "desc" in h or "name" in h or "label" in h) and "desc" not in cols:
                cols["desc"] = i + 1
        if {"code", "type", "desc"}.issubset(cols.keys()):
            return row_idx, cols
    raise ValueError("Could not detect header row in first 10 rows")


def main():
    print("=" * 60)
    print(" v0.4.4.1 hotfix - Add PL (Overhead Powerline) to ODOT catalog")
    print("=" * 60)
    xlsx = find_xlsx()
    wb = load_workbook(xlsx)
    ws = wb.active
    try:
        header_row, cols = detect_header_row(ws)
        print(f"[OK] Detected header at row {header_row}; columns: {cols}")
    except ValueError as e:
        print(f"[ERROR] {e}")
        sys.exit(1)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code_val = row[cols["code"] - 1] if len(row) >= cols["code"] else None
        if code_val and str(code_val).strip().upper() == NEW_CODE:
            print(f"[OK] PL already exists in {xlsx.name}. Nothing to do.")
            return
    last_row = ws.max_row
    target_row = last_row + 1
    ws.cell(row=target_row, column=cols["code"], value=NEW_CODE)
    ws.cell(row=target_row, column=cols["type"], value=NEW_TYPE)
    ws.cell(row=target_row, column=cols["desc"], value=NEW_DESC)
    wb.save(xlsx)
    print(f"[OK] Added {NEW_CODE} ({NEW_TYPE}) - '{NEW_DESC}' at row {target_row}")
    print(f"[OK] Saved {xlsx}")
    print()
    print("NEXT: The apply script will now reseed the translation map.")


if __name__ == "__main__":
    main()
