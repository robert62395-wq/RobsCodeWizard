"""Standalone translation_map.json seeder for Rob's Code Wizard.

Reads VDT_CODES.xlsx and ODOT_CODES.xlsx and produces a populated
app/data/translation_map.json. Skips title rows automatically by scanning
the first 10 rows looking for a real header row.
"""
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed. Run:")
    print("    .venv\\Scripts\\python -m pip install openpyxl rapidfuzz")
    sys.exit(1)

try:
    from rapidfuzz import fuzz
except ImportError:
    print("ERROR: rapidfuzz is not installed. Run:")
    print("    .venv\\Scripts\\python -m pip install rapidfuzz")
    sys.exit(1)

PROJECT_ROOT = Path(__file__).parent
VDT_XLSX_CANDIDATES = [
    PROJECT_ROOT / "app" / "data" / "VDT_CODES.xlsx",
    PROJECT_ROOT / "resources" / "VDT_CODES.xlsx",
    PROJECT_ROOT / "VDT_CODES.xlsx",
]
ODOT_XLSX_CANDIDATES = [
    PROJECT_ROOT / "app" / "data" / "ODOT_CODES.xlsx",
    PROJECT_ROOT / "resources" / "ODOT_CODES.xlsx",
    PROJECT_ROOT / "ODOT_CODES.xlsx",
]
MAP_PATH = PROJECT_ROOT / "app" / "data" / "translation_map.json"

COLUMN_HINTS = {
    "code":        ["code", "field code", "abbr"],
    "type":        ["type", "category", "class"],
    "description": ["description", "desc", "name", "label"],
}

EXACT_THRESHOLD = 95
BEST_GUESS_THRESHOLD = 70


def find_xlsx(candidates, label):
    for p in candidates:
        if p.exists():
            print(f"[OK] Found {label}: {p}")
            return p
    print(f"[ERROR] Could not find {label}.xlsx. Looked in:")
    for p in candidates:
        print(f"    {p}")
    sys.exit(1)


def detect_columns(header_row):
    mapping = {}
    lower_headers = [str(h or "").strip().lower() for h in header_row]
    for logical, hints in COLUMN_HINTS.items():
        for i, header in enumerate(lower_headers):
            if any(hint in header for hint in hints):
                mapping[logical] = i
                break
    missing = set(COLUMN_HINTS) - set(mapping)
    if missing:
        raise ValueError(f"Could not detect columns for: {missing}. Headers: {lower_headers}")
    return mapping


def load_codes(xlsx_path, label):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"[ERROR] {label} is empty.")
        sys.exit(1)
    header_row_index = None
    cols = None
    for i, candidate in enumerate(rows[:10]):
        lower = [str(h or "").strip().lower() for h in candidate]
        non_empty = sum(1 for c in lower if c)
        if non_empty >= 3 and any("code" in c for c in lower):
            try:
                cols = detect_columns(candidate)
                header_row_index = i
                print(f"[OK] {label}: detected header at row {i+1}")
                break
            except ValueError:
                continue
    if header_row_index is None:
        print(f"[ERROR] No header row found in first 10 rows of {label}.")
        sys.exit(1)
    codes = []
    for r in rows[header_row_index + 1:]:
        if not r or not r[cols["code"]]:
            continue
        codes.append({
            "code":        str(r[cols["code"]]).strip(),
            "type":        str(r[cols["type"]] or "").strip(),
            "description": str(r[cols["description"]] or "").strip(),
        })
    print(f"[OK] Loaded {len(codes)} codes from {label}")
    return codes


def score_pair(vdt_row, odot_row):
    if vdt_row["type"] != odot_row["type"]:
        return 0
    return fuzz.token_set_ratio(vdt_row["description"], odot_row["description"])


def best_match(vdt_row, candidates):
    best, best_score = None, 0
    for c in candidates:
        s = score_pair(vdt_row, c)
        if s > best_score:
            best, best_score = c, s
    return best, best_score


def build_entry(vdt_row, odot_row, score):
    if score >= EXACT_THRESHOLD:
        confidence = "exact"
    elif score >= BEST_GUESS_THRESHOLD:
        confidence = "best-guess"
    else:
        confidence = "unmatched"
        odot_row = None
    return {
        "id": f"vdt-{vdt_row['code']}->odot-{odot_row['code'] if odot_row else 'NONE'}",
        "vdt": vdt_row,
        "odot": odot_row,
        "confidence": confidence,
        "match_basis": ["description", "type"] if odot_row else [],
        "score": round(score / 100, 2),
        "user_override": False,
        "notes": "",
    }


def main():
    print("=" * 60)
    print(" Rob's Code Wizard - Standalone Translation Map Seeder")
    print("=" * 60)
    vdt_path = find_xlsx(VDT_XLSX_CANDIDATES, "VDT_CODES")
    odot_path = find_xlsx(ODOT_XLSX_CANDIDATES, "ODOT_CODES")
    vdt_rows = load_codes(vdt_path, "VDT_CODES.xlsx")
    odot_rows = load_codes(odot_path, "ODOT_CODES.xlsx")

    entries = []
    matched_odot_codes = set()
    print("\nMatching VDT -> ODOT...")
    for v in vdt_rows:
        candidates = [o for o in odot_rows if o["type"] == v["type"]]
        match, score = best_match(v, candidates)
        entry = build_entry(v, match, score)
        entries.append(entry)
        if match and score >= BEST_GUESS_THRESHOLD:
            matched_odot_codes.add(match["code"])

    print("Adding orphan ODOT codes...")
    for o in odot_rows:
        if o["code"] not in matched_odot_codes:
            entries.append({
                "id": f"vdt-NONE->odot-{o['code']}",
                "vdt": None,
                "odot": o,
                "confidence": "unmatched",
                "match_basis": [],
                "score": 0.0,
                "user_override": False,
                "notes": "No VDT counterpart auto-detected.",
            })

    entries.sort(key=lambda e: (
        e["vdt"]["code"] if e["vdt"] else "zzz",
        e["odot"]["code"] if e["odot"] else "zzz",
    ))

    data = {
        "schema_version": "1.0",
        "map_version": "0.4.1-standalone-seed",
        "generated": datetime.now(timezone.utc).isoformat(),
        "entries": entries,
    }
    MAP_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(MAP_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    summary = {
        "total": len(entries),
        "exact": sum(1 for e in entries if e["confidence"] == "exact"),
        "best_guess": sum(1 for e in entries if e["confidence"] == "best-guess"),
        "unmatched": sum(1 for e in entries if e["confidence"] == "unmatched"),
    }
    print()
    print("=" * 60)
    print(f"Wrote {MAP_PATH}")
    print(f"  Total entries:    {summary['total']}")
    print(f"  Exact:            {summary['exact']}")
    print(f"  Best-guess:       {summary['best_guess']}")
    print(f"  Unmatched:        {summary['unmatched']}")
    print("=" * 60)
    print("Restart the app. Translation tab should now show entries.")


if __name__ == "__main__":
    main()
