from pathlib import Path
from openpyxl import load_workbook

ODOT_XLSX = Path(__file__).resolve().parents[1] / "resources" / "ODOT_CODES.xlsx"
EXPECTED = ["Category","Code","Type","DTM","Description","AttributesSchema","SymbolCell","ShotLocation"]


def _load():
    wb = load_workbook(ODOT_XLSX, read_only=True, data_only=True)
    return list(wb.active.iter_rows(values_only=True))


def test_exists():
    assert ODOT_XLSX.exists()


def test_headers():
    rows = _load()
    assert [c for c in rows[1]] == EXPECTED


def test_codes_count_and_spot_checks():
    rows = _load()
    codes = {str(r[1]).strip() for r in rows[2:] if r and r[1]}
    assert len(codes) >= 300
    for expected in ("PP","FH","EP","BLD","CB-7","MH-3-SAN","BLD_2D","AERTAR"):
        assert expected in codes


def test_types_classification():
    rows = _load()
    types = {str(r[2]).strip() for r in rows[2:] if r and r[2]}
    assert "Linework" in types and "Symbol" in types and "Point" in types


def test_dtm_flags():
    rows = _load()
    dtms = {str(r[3]).strip() for r in rows[2:] if r and r[3]}
    assert {"DNI","BREAK","SPOT","VOID"} <= dtms
