"""Linework fix tests (Phase 4 extended)."""
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from app.services.linework_fix import (
    run_linework_fix, audit_with_locations, suggest_fix,
    export_audit_txt, export_audit_xlsx,
)
from app.services.catalog_loader import load_catalog


def _rows(*descs):
    return [{"P": str(i), "N": "0", "E": "0", "Z": "100", "D": d}
            for i, d in enumerate(descs, 1)]


# --- legacy tests (must still pass) ---

def test_balanced_vdt():
    vdt = load_catalog("vdt")
    a = run_linework_fix(_rows("EPA B", "EPA E"), vdt)
    assert a["begins"]["EPA"] == 1 and a["ends"]["EPA"] == 1
    assert a["unbalanced"] == {}


def test_unbalanced_vdt():
    vdt = load_catalog("vdt")
    a = run_linework_fix(_rows("EPA B", "EPA B"), vdt)
    assert a["unbalanced"] == {"EPA": 2}


def test_multi_codes_in_one_row():
    vdt = load_catalog("vdt")
    a = run_linework_fix(_rows("EPA E EPC B"), vdt)
    assert a["ends"]["EPA"] == 1 and a["begins"]["EPC"] == 1


def test_odot_curve_toggle():
    odot = load_catalog("odot")
    a = run_linework_fix(_rows("EP OC*", "EP OC*"), odot)
    assert a["begins"].get("EP") == 1 and a["ends"].get("EP") == 1
    assert a["unbalanced"] == {}


def test_no_grammar_returns_empty():
    a = run_linework_fix(_rows("EPA B"), None)
    assert a == {"begins": {}, "ends": {}, "unbalanced": {}, "total_commands": 0}


# --- Phase 4 new tests ---

def test_audit_with_locations_tracks_row_indices():
    vdt = load_catalog("vdt")
    a = audit_with_locations(_rows("EPA B", "EPA B"), vdt)
    assert a["row_indices"]["EPA"] == [0, 1]


def test_suggest_fix_missing_end_vdt():
    vdt = load_catalog("vdt")
    s = suggest_fix("EPA", +1, vdt)
    assert "end" in s.lower()
    assert "`E`" in s


def test_suggest_fix_missing_begin_vdt():
    vdt = load_catalog("vdt")
    s = suggest_fix("EPA", -1, vdt)
    assert "begin" in s.lower()
    assert "`B`" in s


def test_suggest_fix_balanced():
    vdt = load_catalog("vdt")
    s = suggest_fix("EPA", 0, vdt)
    assert "balanced" in s.lower()


def test_suggest_fix_no_grammar():
    s = suggest_fix("EPA", +2, None)
    assert "unbalanced" in s.lower()


def test_export_audit_txt():
    vdt = load_catalog("vdt")
    audit = audit_with_locations(_rows("EPA B", "EPA E"), vdt)
    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "audit.txt"
        export_audit_txt(audit, out)
        text = out.read_text(encoding="utf-8")
        assert "Linework Fix Audit" in text
        assert "Per-Code Summary" in text


def test_export_audit_xlsx_has_both_sheets():
    vdt = load_catalog("vdt")
    audit = audit_with_locations(_rows("EPA B", "EPA B"), vdt)
    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "audit.xlsx"
        export_audit_xlsx(audit, out, vdt)
        wb = load_workbook(out)
        assert "Summary" in wb.sheetnames
        assert "Per-Code" in wb.sheetnames
        per_code = wb["Per-Code"]
        # Header row + at least one data row
        assert per_code.max_row >= 2
        # Suggestion column has content for unbalanced EPA
        headers = [c.value for c in per_code[1]]
        sug_col = headers.index("Suggestion") + 1
        sugs = [per_code.cell(row=r, column=sug_col).value
                for r in range(2, per_code.max_row + 1)]
        assert any(s and "end" in s.lower() for s in sugs)