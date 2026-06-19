import tempfile
from pathlib import Path
from openpyxl import load_workbook
from app.services.report_exporter import build_report, export_xlsx, export_txt, export_report

ROWS = [{"P":"1","N":"1","E":"1","Z":"100","D":"EPA"},
        {"P":"2","N":"1","E":"1","Z":"0","D":"ZZZ"}]
RESULTS = [{"valid":True,"issues":[]},
           {"valid":False,"issues":["Zero elevation","Unknown code(s): ZZZ"]}]
SUGG = ["", "EPA"]


def test_build_report_counts():
    r = build_report(ROWS, RESULTS, SUGG)
    assert r["summary"]["total_rows"] == 2
    assert r["summary"]["valid_rows"] == 1


def test_export_xlsx_writes_sheets():
    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "r.xlsx"
        export_xlsx(build_report(ROWS, RESULTS, SUGG), out)
        wb = load_workbook(out)
        assert "Summary" in wb.sheetnames and "Rows" in wb.sheetnames


def test_export_txt():
    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "r.txt"
        export_txt(build_report(ROWS, RESULTS, SUGG), out)
        assert "Validation Report" in out.read_text(encoding="utf-8")


def test_export_report_dispatches():
    with tempfile.TemporaryDirectory() as td:
        x = Path(td) / "r.xlsx"; t = Path(td) / "r.txt"
        export_report(ROWS, RESULTS, SUGG, x); export_report(ROWS, RESULTS, SUGG, t)
        assert x.exists() and t.exists()
