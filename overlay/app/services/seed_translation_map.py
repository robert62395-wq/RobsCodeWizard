"""Reseed translation_map.json from XLSX (v0.5.1.1)."""
from __future__ import annotations
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from rapidfuzz import fuzz

from app.services import translation_map as tm

PROJECT_ROOT = Path(__file__).parent.parent.parent
VDT_XLSX_CANDIDATES = [
    PROJECT_ROOT / "app" / "data" / "VDT_CODES.xlsx",
    PROJECT_ROOT / "resources" / "VDT_CODES.xlsx",
    PROJECT_ROOT / "VDT_CODES.xlsx",
]
ODOT_XLSX_CANDIDATES = [
    PROJECT_ROOT / "app" / "data" / "ODOT_CODES.xlsx",
    PROJECT_ROOT / "resources" / "ODOT_CODES.xlsx",
    PROJECT_ROOT / "ODOT_CODES.xlsx",
]

COLUMN_HINTS = {
    "code":        ["code", "field code", "abbr"],
    "type":        ["type", "category", "class"],
    "description": ["description", "desc", "name", "label"],
}

EXACT_THRESHOLD = 95
BEST_GUESS_THRESHOLD = 70
MAP_VERSION = "0.5.1.1"


def _find_xlsx(candidates, label):
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(f"Could not find {label}.xlsx")


def _detect_columns(header_row):
    mapping = {}
    lower = [str(h or "").strip().lower() for h in header_row]
    for logical, hints in COLUMN_HINTS.items():
        for i, h in enumerate(lower):
            if any(hint in h for hint in hints):
                mapping[logical] = i
                break
    missing = set(COLUMN_HINTS) - set(mapping)
    if missing:
        raise ValueError(f"Could not detect columns for: {missing}")
    return mapping


def _load_codes(xlsx_path, label):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{label} is empty")
    header_row_index = None
    cols = None
    for i, candidate in enumerate(rows[:10]):
        lower = [str(h or "").strip().lower() for h in candidate]
        non_empty = sum(1 for c in lower if c)
        if non_empty >= 3 and any("code" in c for c in lower):
            try:
                cols = _detect_columns(candidate)
                header_row_index = i
                break
            except ValueError:
                continue
    if header_row_index is None:
        raise ValueError(f"No header row in {label}")
    codes = []
    for r in rows[header_row_index + 1:]:
        if not r or not r[cols["code"]]:
            continue
        codes.append({
            "code":        str(r[cols["code"]]).strip(),
            "type":        str(r[cols["type"]] or "").strip(),
            "description": str(r[cols["description"]] or "").strip(),
        })
    return codes


def _score(v, o):
    if v["type"] != o["type"]:
        return 0
    return fuzz.token_set_ratio(v["description"], o["description"])


def _best_match(v, odot_rows):
    best, best_score = None, 0
    for o in odot_rows:
        s = _score(v, o)
        if s > best_score:
            best, best_score = o, s
    return best, best_score


def _build_entry(v, o, score):
    if score >= EXACT_THRESHOLD:
        conf = "exact"
    elif score >= BEST_GUESS_THRESHOLD:
        conf = "best-guess"
    else:
        conf = "unmatched"
        o = None
    return {
        "id": f"vdt-{v['code']}->odot-{o['code'] if o else 'NONE'}",
        "vdt": v, "odot": o,
        "confidence": conf,
        "match_basis": ["description", "type"] if o else [],
        "score": round(score / 100, 2),
        "user_override": False,
        "notes": "",
    }


def seed(force=False):
    if tm.MAP_PATH.exists() and not force:
        existing = tm.load()
        if existing.get("entries"):
            return existing

    vdt_path = _find_xlsx(VDT_XLSX_CANDIDATES, "VDT_CODES")
    odot_path = _find_xlsx(ODOT_XLSX_CANDIDATES, "ODOT_CODES")
    vdt_rows = _load_codes(vdt_path, "VDT_CODES.xlsx")
    odot_rows = _load_codes(odot_path, "ODOT_CODES.xlsx")

    entries = []
    matched_odot = set()
    for v in vdt_rows:
        candidates = [o for o in odot_rows if o["type"] == v["type"]]
        match, score = _best_match(v, candidates)
        entry = _build_entry(v, match, score)
        entries.append(entry)
        if match and score >= BEST_GUESS_THRESHOLD:
            matched_odot.add(match["code"])

    for o in odot_rows:
        if o["code"] not in matched_odot:
            entries.append({
                "id": f"vdt-NONE->odot-{o['code']}",
                "vdt": None, "odot": o,
                "confidence": "unmatched",
                "match_basis": [],
                "score": 0.0,
                "user_override": False,
                "notes": "No VDT counterpart auto-detected.",
            })

    entries.sort(key=lambda e: (
        e["vdt"]["code"] if e["vdt"] else "zzz",
        e["odot"]["code"] if e["odot"] else "zzz",
    ))

    data = {
        "schema_version": tm.SCHEMA_VERSION,
        "map_version": MAP_VERSION,
        "generated": datetime.now(timezone.utc).isoformat(),
        "entries": entries,
    }
    tm.save(data)
    return data


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--force", action="store_true")
    args = p.parse_args()
    result = seed(force=args.force)
    print(f"[OK] Seeded {len(result['entries'])} entries")
